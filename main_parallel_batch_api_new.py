import requests
import os
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm
import pandas as pd
from typing import List, Dict

# --------------------------
# 配置项
# --------------------------
API_URL = "http://10.136.234.255:8080/process_image"
API_URL = "http://10.136.234.255:8081/process_image"

MAX_WORKERS = 1  # 线程池大小
REQUEST_TIMEOUT = 300  # 请求超时时间（秒）
dir_pre = "/Users/zhipeng/Win10/LocalOneDrive/Gitee/Multi_agent_image_tagging"
dir_pre = "/workspace/work/zhipeng16/git/Multi_agent_image_tagging"
IMAGE_FOLDER = dir_pre + '/无他图片标签测试图2'  # 图片文件夹路径
IMAGE_FOLDER = dir_pre + '/badcase' 
OUTPUT_EXCEL = dir_pre + "/主体类型.xlsx"
OUTPUT_EXCEL2 = dir_pre + "/主体类型2.xlsx"
PREFIX_TO_REMOVE = "/Users/zhipeng/Win10/LocalOneDrive/Gitee"  # 路径简化前缀
PREFIX_ONLINE = "/workspace/work/zhipeng16/git"  # 路径简化前缀
# --------------------------
# 单张图片接口调用函数
# --------------------------
def call_image_api(img_path: str, API_URL = API_URL) -> Dict:
    """
    调用图片标签接口，返回单张图片的处理结果
    """
    # 构造请求体
    task_id = str(uuid.uuid4())  # 生成唯一task_id
    request_data = {
        "image_info": img_path,
        "task_id": task_id
    }
    
    # 构造请求头
    headers = {
        "Content-Type": "application/json"
    }
    
    try:
        # 发送POST请求
        response = requests.post(
            url=API_URL,
            json=request_data,
            headers=headers,
            timeout=REQUEST_TIMEOUT
        )
        response.raise_for_status()  # 抛出HTTP状态码异常
        
        # 解析响应结果
        response_json = response.json()
        if response_json.get("code") == 200:
            return response_json.get("res", {})
        else:
            return {
                "image_info": img_path,
                "final_labels": [],
                "total_labels_count": 0,
                "elapsed_time": 0.0,
                "token_cost": 0.0,
                "status": "failed",
                "error": f"接口返回错误码：{response_json.get('code')}，详情：{response_json.get('detail', '')}"
            }
    
    except requests.exceptions.RequestException as e:
        # 捕获请求相关异常
        return {
            "image_info": img_path,
            "final_labels": [],
            "total_labels_count": 0,
            "elapsed_time": 0.0,
            "token_cost": 0.0,
            "status": "failed",
            "error": f"请求失败：{str(e)}"
        }
    except Exception as e:
        # 捕获其他异常
        return {
            "image_info": img_path,
            "final_labels": [],
            "total_labels_count": 0,
            "elapsed_time": 0.0,
            "token_cost": 0.0,
            "status": "failed",
            "error": f"未知错误：{str(e)}"
        }

# --------------------------
# 批量接口调用函数
# --------------------------
def batch_call_image_api_new(image_paths: List[str], API_URL = API_URL) -> List[Dict]:
    """
    批量调用图片标签接口，使用线程池并发处理
    """
    results = []
    results2 = []

    for img_path in image_paths:
        try:
            result = call_image_api(img_path, API_URL)
            # result2 = call_image_api(img_path,API_URL2)
            results.append(result)
            # results2.append(result2)
        except Exception as e:

            error_result = {
                "image_info": img_path,
                "final_labels": [],
                "total_labels_count": 0,
                "elapsed_time": 0.0,
                "token_cost": 0.0,
                "status": "failed",
                "error": f"线程执行异常：{str(e)}"
            }
            results.append(error_result)

    return results

def batch_call_image_api(image_paths: List[str]) -> List[Dict]:
    """
    批量调用图片标签接口，使用线程池并发处理
    """
    results = []
    
    # 使用线程池并发执行
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        # 提交所有任务
        future_to_img_path = {
            executor.submit(call_image_api, img_path): img_path 
            for img_path in image_paths
        }
        
        # 遍历完成的任务，获取结果
        for future in tqdm(as_completed(future_to_img_path), total=len(future_to_img_path), desc="批量调用接口"):
            img_path = future_to_img_path[future]
            try:
                result = future.result()
                results.append(result)
            except Exception as e:
                # 捕获线程执行异常
                error_result = {
                    "image_info": img_path,
                    "final_labels": [],
                    "total_labels_count": 0,
                    "elapsed_time": 0.0,
                    "token_cost": 0.0,
                    "status": "failed",
                    "error": f"线程执行异常：{str(e)}"
                }
                results.append(error_result)
    
    return results

# --------------------------
# 结果保存到Excel函数（包含标签包含性判断）
# --------------------------
def save_api_results_to_excel(results: List[Dict], output_file: str = OUTPUT_EXCEL) -> str:
    """
    将接口返回结果保存到Excel，包含标签包含性判断
    """
    excel_data = []
    prefix = PREFIX_TO_REMOVE
    
    for idx, result in enumerate(results, 1):
        img_path = result.get("image_info", "")
        # 简化路径显示
        img_relative_path = img_path[len(prefix):] if img_path.startswith(prefix) else img_path
        filename = os.path.basename(img_path)
        file_dir = os.path.dirname(img_path)
        
        # 第1列：ID + 路径名
        id_path = f"{idx:03d} - {img_relative_path}"
        
        # 第2列：预测标签（用|分隔）
        predicted_labels = result.get("final_labels", [])
        predicted_labels_str = "|".join(predicted_labels) if predicted_labels else ""
        
        # 第3列：路径标签（提取逻辑沿用原有批量测试代码）
        # 提取文件夹名标签（去掉数字+顿号）
        dir_name = os.path.basename(file_dir)
        dir_label = dir_name.split("、")[-1] if "、" in dir_name else dir_name
        
        # 提取文件名标签（去掉数字）
        file_prefix = filename.split(".")[0] if "." in filename else filename
        file_label = "".join([c for c in file_prefix if not c.isdigit()])
        
        # 拼接路径标签
        path_label = f"{dir_label}-{file_label}"
        
        # 第4列：是否包含（判断预测标签是否包含路径标签核心词）
        is_include = "N/A"
        if predicted_labels:
            predicted_str = "|".join(predicted_labels).lower()
            core_dir_label = dir_label.lower()
            is_include = "是" if core_dir_label in predicted_str else "否"
        
        # 组装行数据
        row = {
            "ID_路径名": id_path,
            "预测标签": predicted_labels_str,
            "路径标签": path_label,
            "是否包含": is_include,
            "耗时(s)": round(result.get("elapsed_time", 0), 2),
            "Token耗费(¥)": round(result.get("token_cost", 0), 4),
            "标签总数": result.get("total_labels_count", 0),
            "处理状态": result.get("status", "unknown"),
            "错误信息": result.get("error", "")
        }
        excel_data.append(row)
    
    # 生成Excel文件
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 第一个Sheet：详细结果
        df_detail = pd.DataFrame(excel_data)
        df_detail.to_excel(writer, sheet_name='标签对比分析', index=False)
        
        # 第二个Sheet：统计汇总
        total_count = len(results)
        success_count = len([r for r in results if r.get("status") == "success"])
        success_rate = round(success_count / total_count * 100, 1) if total_count > 0 else 0
        total_elapsed = round(sum(r.get("elapsed_time", 0) for r in results), 1)
        total_cost = round(sum(r.get("token_cost", 0) for r in results), 4)
        avg_cost = round(total_cost / total_count, 4) if total_count > 0 else 0
        
        summary_data = {
            "统计项": ["总图片数", "成功数", "成功率(%)", "总耗时(s)", "总成本(¥)", "平均成本/图(¥)"],
            "数值": [
                total_count,
                success_count,
                success_rate,
                total_elapsed,
                total_cost,
                avg_cost
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='统计汇总', index=False)
        
        # 设置Excel样式（可选，保持原有格式）
        worksheet = writer.sheets['标签对比分析']
        # 列宽设置
        column_widths = {
            'A': 30, 'B': 60, 'C': 20, 'D': 10,
            'E': 10, 'F': 12, 'G': 10, 'H': 10, 'I': 30
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 表头样式（可选）
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # "是否包含"列条件格式
        for row_idx in range(2, len(excel_data) + 2):
            cell = worksheet[f'D{row_idx}']
            if cell.value == "是":
                cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            elif cell.value == "否":
                cell.fill = PatternFill(start_color="F75B5B", end_color="F75B5B", fill_type="solid")
    
    # 打印统计信息
    print("\n📊 批量接口测试统计汇总：")
    print(f"   总图片数：{total_count}")
    print(f"   成功数：{success_count} | 成功率：{success_rate}%")
    print(f"   总耗时：{total_elapsed}s")
    print(f"   总成本：¥{total_cost:.4f}")
    print(f"   平均成本/图：¥{avg_cost:.4f}")
    print(f"\n💾 结果已保存至Excel：{output_file}")
    
    return output_file

# --------------------------
# 主函数：扫描图片 + 批量调用 + 保存Excel
# --------------------------
if __name__ == "__main__":
    # 1. 扫描指定文件夹下的所有图片
    image_paths = []
    supported_formats = ('.png', '.jpg', '.jpeg')
    for root, _, files in os.walk(IMAGE_FOLDER):
        for file in files:
            if file.lower().endswith(supported_formats):
                new_root = root.replace(PREFIX_TO_REMOVE,PREFIX_ONLINE)
                img_full_path = os.path.join(new_root, file)
                image_paths.append(img_full_path)
    
    # 打印扫描结果
    print(f"📁 扫描完成：在 {IMAGE_FOLDER} 下发现 {len(image_paths)} 张图片")
    if not image_paths:
        print("⚠️ 未发现任何图片，程序退出")
        exit(1)
    
    # 2. 批量调用接口
    print("\n🚀 开始批量调用图片标签接口...")
    batch_results = batch_call_image_api_new(image_paths, API_URL)
    
    # 3. 保存结果到Excel
    print("\n📝 开始保存结果到Excel...")
    save_api_results_to_excel(batch_results, OUTPUT_EXCEL)
    
    print("\n✅ 所有任务执行完成！")